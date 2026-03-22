"""
Extraction helpers ported from sber_hack_2026 ConvertService (csv.Sniffer DictReader,
openpyxl streaming, pypdf text, img2table + OCR variants). All imports are optional
where noted; failures fall back to file_parser defaults.
"""

from __future__ import annotations

import csv
import io
import re
from typing import Any, Dict, List, Optional, Tuple

# --- CSV / XLSX (always use stdlib + openpyxl; openpyxl is in requirements) ---


def _stringify_cell(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, bool):
        return "true" if value else "false"
    return str(value).strip()


def _decode_text_bytes(data: bytes) -> str:
    for encoding in ("utf-8-sig", "utf-8", "cp1251"):
        try:
            return data.decode(encoding)
        except UnicodeDecodeError:
            continue
    return data.decode("utf-8", errors="replace")


def parse_csv_dict_rows(file_bytes: bytes) -> Tuple[List[Dict[str, str]], str]:
    """
    Port of sber_hack_2026 ConvertService._parse_csv_rows.
    Returns (rows, delimiter_char).
    """
    decoded = _decode_text_bytes(file_bytes)
    stream = io.StringIO(decoded)
    sample = decoded[:4096]
    try:
        dialect = csv.Sniffer().sniff(sample, delimiters=";,|\t,")
    except csv.Error:
        dialect = csv.excel
        dialect.delimiter = ";"

    delimiter = str(getattr(dialect, "delimiter", ";") or ";")
    stream.seek(0)
    reader = csv.DictReader(stream, dialect=dialect)
    headers = [h.strip() for h in (reader.fieldnames or []) if h and h.strip()]
    if not headers:
        raise ValueError("CSV headers are empty")

    rows: List[Dict[str, str]] = []
    for row in reader:
        clean_row = {str(k).strip(): _stringify_cell(v) for k, v in row.items() if k}
        if any(value != "" for value in clean_row.values()):
            rows.append(clean_row)
    if not rows:
        raise ValueError("CSV has no data rows")
    return rows, delimiter


def parse_xlsx_dict_rows(file_bytes: bytes, *, max_rows: Optional[int] = None) -> List[Dict[str, str]]:
    """Port of sber_hack_2026 ConvertService._parse_xlsx_rows with optional row cap."""
    from openpyxl import load_workbook

    workbook = load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
    try:
        sheet = workbook.active
        rows_iter = sheet.iter_rows(values_only=True)

        headers: List[str] = []
        for raw_row in rows_iter:
            candidate = [_stringify_cell(cell) for cell in raw_row]
            if any(candidate):
                headers = [value.strip() for value in candidate if value.strip()]
                break
        if not headers:
            raise ValueError("XLSX headers are empty")

        parsed_rows: List[Dict[str, str]] = []
        for raw_row in rows_iter:
            values = [_stringify_cell(cell) for cell in raw_row]
            row_map: Dict[str, str] = {}
            has_value = False
            for index, header in enumerate(headers):
                value = values[index].strip() if index < len(values) else ""
                row_map[header] = value
                if value:
                    has_value = True
            if has_value:
                parsed_rows.append(row_map)
            if max_rows is not None and len(parsed_rows) >= max_rows:
                break
        return parsed_rows
    finally:
        workbook.close()


# --- PDF (pypdf preferred) ---


def extract_pdf_text_pypdf(file_bytes: bytes) -> str:
    try:
        from pypdf import PdfReader
    except ImportError:
        return ""

    try:
        reader = PdfReader(io.BytesIO(file_bytes))
    except Exception:
        return ""

    text_parts: List[str] = []
    try:
        for page in reader.pages:
            text = page.extract_text() or ""
            if text.strip():
                text_parts.append(text)
    except Exception:
        return ""
    return _normalize_whitespace("\n".join(text_parts))


def _normalize_whitespace(text: str) -> str:
    return re.sub(r"\s+", " ", text).strip()


def _capture_group(text: str, pattern: str) -> str:
    match = re.search(pattern, text, flags=re.IGNORECASE | re.DOTALL)
    if not match:
        return ""
    return _normalize_whitespace(match.group(1))


def _is_marked_option(text: str, phrase: str) -> bool:
    lowered = text.lower()
    phrase_lower = phrase.lower()
    index = lowered.find(phrase_lower)
    if index == -1:
        return False
    window_start = max(0, index - 60)
    window = text[window_start:index]
    return bool(re.search(r"\bX\b", window, flags=re.IGNORECASE))


def extract_fatca_row_from_text(text: str) -> Dict[str, str]:
    """Port of sber _extract_fatca_row — enriches CRM-style forms when markers match."""
    import json

    organization_name = _capture_group(text, r"Наименование организации\s+(.+?)\s+ИНН/КИО")
    inn_or_kio = _capture_group(text, r"ИНН/КИО\s+(\d{10,12})")

    is_resident_rf = ""
    if _is_marked_option(text, "Не являюсь налоговым резидентом ни в одном"):
        is_resident_rf = "NOWHERE"
    elif _is_marked_option(text, "ДА, является налоговым резидентом только в РФ"):
        is_resident_rf = "YES"
    elif _is_marked_option(text, "НЕТ, является налоговым резидентом"):
        is_resident_rf = "NO"

    is_tax_residency_only_rf = ""
    if _is_marked_option(text, "все контролирующие лица являются налоговыми резидентами только в РФ"):
        is_tax_residency_only_rf = "YES"
    elif _is_marked_option(text, "среди контролирующих лиц есть налоговые резиденты иностранных юрисдикций"):
        is_tax_residency_only_rf = "NO"

    options: List[str] = []
    option_rules = [
        ("IS_DISREGARDED_ENTITY", "disregarded entity"),
        ("IS_FATCA_FOREIGN_INSTITUTE", "Иностранным финансовым институтом"),
        ("TEN_OR_MORE_PERCENT_IN_USA", "Более 10% акций"),
        ("STATEMENTS_NOT_APPILCABLE", "данные утверждения не применимы"),
    ]
    for code, phrase in option_rules:
        if _is_marked_option(text, phrase):
            options.append(code)

    return {
        "organizationName": organization_name,
        "innOrKio": inn_or_kio,
        "isResidentRF": is_resident_rf,
        "isTaxResidencyOnlyRF": is_tax_residency_only_rf,
        "fatcaBeneficiaryOptionList": json.dumps(options, ensure_ascii=False),
        "sourceText": text[:2000],
    }


def looks_like_fatca_text(text: str) -> bool:
    normalized = text.lower()
    markers = [
        "инн/кио",
        "fatca",
        "выгодоприобретател",
        "налоговым резидентом",
    ]
    score = sum(1 for marker in markers if marker in normalized)
    return score >= 2


# --- Images: img2table + PIL variants (optional) ---


def _normalize_table_header(value: str) -> str:
    header = _normalize_whitespace(value)
    header = header.strip(":;,.")
    header = header.replace(" ", "_").replace("-", "_")
    header = re.sub(r"_{2,}", "_", header)
    return header


def _deduplicate_headers(headers: List[str]) -> List[str]:
    result: List[str] = []
    counts: Dict[str, int] = {}
    for raw_header in headers:
        base = raw_header.strip()
        if not base:
            result.append("")
            continue
        index = counts.get(base, 0) + 1
        counts[base] = index
        result.append(base if index == 1 else f"{base}_{index}")
    return result


def _score_extracted_table(rows: List[Dict[str, str]]) -> float:
    if not rows:
        return 0.0
    headers = list(rows[0].keys())
    if not headers:
        return 0.0
    valid_header_count = sum(1 for header in headers if re.match(r"^[A-Za-z_][A-Za-z0-9_]*$", header))
    header_score = valid_header_count / max(len(headers), 1)
    non_empty_values = 0
    total_values = 0
    for row in rows[:20]:
        for value in row.values():
            total_values += 1
            if value.strip():
                non_empty_values += 1
    fill_score = non_empty_values / max(total_values, 1)
    return (header_score * 0.7) + (fill_score * 0.3)


def _image_variants(file_bytes: bytes) -> List[bytes]:
    try:
        from PIL import Image, ImageEnhance, ImageFilter, ImageOps
    except ImportError:
        return [file_bytes]

    variants: List[bytes] = [file_bytes]
    try:
        source = Image.open(io.BytesIO(file_bytes))
        gray = ImageOps.grayscale(source)
        enhanced = ImageEnhance.Contrast(gray).enhance(2.2)
        sharpened = enhanced.filter(ImageFilter.SHARPEN)
        upscaled = sharpened.resize((sharpened.width * 2, sharpened.height * 2), Image.Resampling.LANCZOS)
        binary = upscaled.point(lambda p: 255 if p > 160 else 0)
        buffer = io.BytesIO()
        binary.save(buffer, format="PNG")
        variants.append(buffer.getvalue())
    except Exception:
        return variants
    return variants


def extract_image_table_rows_best_effort(file_bytes: bytes) -> List[Dict[str, str]]:
    """Port of sber _parse_image_table_rows; returns [] if deps missing or no table."""
    try:
        from img2table.document import Image as Img2TableImage
        from img2table.ocr import TesseractOCR
    except ImportError:
        return []

    best_rows: List[Dict[str, str]] = []
    best_score = -1.0
    ocr = TesseractOCR(n_threads=1, lang="rus+eng")
    for variant_bytes in _image_variants(file_bytes):
        rows: List[Dict[str, str]] = []
        try:
            doc = Img2TableImage(src=io.BytesIO(variant_bytes))
            tables = doc.extract_tables(
                ocr=ocr,
                implicit_rows=True,
                implicit_columns=True,
                borderless_tables=True,
                min_confidence=45,
            )
        except Exception:
            continue

        for table in tables:
            dataframe = getattr(table, "df", None)
            if dataframe is None or dataframe.empty:
                continue
            raw_columns = [_normalize_table_header(str(column)) for column in dataframe.columns]
            if all(not column or column.isdigit() for column in raw_columns):
                header_row = [_normalize_table_header(str(cell)) for cell in dataframe.iloc[0].tolist()]
                data_iter = dataframe.iloc[1:].itertuples(index=False, name=None)
                columns = _deduplicate_headers(header_row)
            else:
                data_iter = dataframe.itertuples(index=False, name=None)
                columns = _deduplicate_headers(raw_columns)

            if not any(columns):
                continue
            for raw_row in data_iter:
                mapped: Dict[str, str] = {}
                has_value = False
                rt = tuple(raw_row)
                for index, header in enumerate(columns):
                    if not header:
                        continue
                    value = _stringify_cell(rt[index] if index < len(rt) else "")
                    mapped[header] = value
                    if value:
                        has_value = True
                if has_value and mapped:
                    rows.append(mapped)

        score = _score_extracted_table(rows)
        if score > best_score:
            best_rows = rows
            best_score = score

    if best_score < 0.35:
        return []
    return best_rows


def ocr_text_from_image_variants(file_bytes: bytes) -> str:
    """Port of sber _ocr_text_from_image (pytesseract + optional RapidOCR)."""
    import pytesseract

    best_text = ""
    try:
        from PIL import Image
    except ImportError:
        return ""

    rapid = None
    try:
        from rapidocr_onnxruntime import RapidOCR

        rapid = RapidOCR()
    except Exception:
        pass

    for variant_bytes in _image_variants(file_bytes):
        try:
            image = Image.open(io.BytesIO(variant_bytes))
            text = pytesseract.image_to_string(image, lang="rus+eng")
        except Exception:
            text = ""
        normalized = _normalize_whitespace(text)
        if len(normalized) > len(best_text):
            best_text = normalized
        if rapid:
            try:
                rapid_result, _ = rapid(variant_bytes)
                fragments: List[str] = []
                for item in rapid_result:
                    if isinstance(item, (list, tuple)) and len(item) >= 2 and isinstance(item[1], str):
                        fragments.append(item[1])
                rapid_text = _normalize_whitespace(" ".join(fragments))
                if len(rapid_text) > len(best_text):
                    best_text = rapid_text
            except Exception:
                continue
    return best_text
